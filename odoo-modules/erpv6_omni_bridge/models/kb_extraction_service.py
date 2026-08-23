import base64
import io
import json
import logging
import re
import time

from odoo import api, models, _
from odoo.exceptions import UserError

from odoo.addons.erpv6_kb.models.kb_knowledge import KB_TYPE_SELECTION

_logger = logging.getLogger(__name__)


class _ChunkTruncatedError(Exception):
    """Interna: il modello ha troncato l'output (finish_reason='length')
    prima di chiudere il JSON. Diversa da un JSONDecodeError generico --
    qui la causa e' nota (output troppo lungo per il blocco) e riprovabile
    dimezzando il blocco, non un problema di formato/contenuto."""


class _ChunkRateLimitedError(Exception):
    """Interna: il provider ha rifiutato la chiamata con un 429 reale
    (rate limit), diverso da un troncamento per lunghezza o da un errore di
    formato -- riprovabile aspettando piu' a lungo, non dimezzando il blocco."""

# Groq tier gratuito impone un limite di 8000 token/minuto su openai/gpt-oss-120b
# (verificato dal vivo: la soglia "Requested" di Groq conta input + max_tokens
# riservato per l'output, non solo l'input effettivo -- una singola chiamata con
# troppo testo va in 413 "tokens per minute", una con testo moderato ma senza
# max_tokens esplicito si vede accorciare l'output a meta' JSON, invalido).
# Per documenti piu' grandi di un chunk, il testo viene spezzato ed elaborato
# a piu' chiamate sequenziali (l'utente ha scelto questa strada gratuita invece
# di un upgrade a pagamento del piano Groq).
CHUNK_MAX_CHARS = 8000
CHUNK_MAX_OUTPUT_TOKENS = 4000
# Pavimento di max_tokens per i sotto-blocchi di uno split: un sotto-blocco
# piccolo non ha bisogno di 4000 token riservati, ma non si scende mai sotto
# questa soglia per lasciare comunque margine a voci KB verbose.
CHUNK_MIN_OUTPUT_TOKENS = 800
# Il limite Groq e' una finestra scorrevole di 60s condivisa da tutte le chiamate
# dell'organizzazione: senza un piano a pagamento non c'e' altro modo per restare
# sotto soglia che aspettare che la chiamata precedente "esca" dalla finestra.
CHUNK_PACING_SECONDS = 65
# Se un blocco e' particolarmente denso (tante righe -> tante voci KB), il JSON
# di output puo' superare CHUNK_MAX_OUTPUT_TOKENS prima di chiudersi: il modello
# lo tronca a meta' stringa (finish_reason='length', non un problema di rate
# limit ne' di formato). SPLIT_MAX_DEPTH limita quante volte un blocco del
# genere viene dimezzato e riprovato prima di arrendersi con l'errore reale
# (evita di dimezzare all'infinito un blocco che fallisce per un'altra ragione).
CHUNK_SPLIT_MAX_DEPTH = 3
# CHUNK_MAX_CHARS da solo non basta: e' un proxy per la dimensione
# dell'INPUT, non dell'OUTPUT. Una riga tabellare compatta (es. export
# .xlsx, "Riga N: campo1 | campo2 | ...") va espansa in un oggetto JSON con
# piu' chiavi nominate per esteso -- l'output di una singola voce puo'
# superare il testo sorgente della riga stessa. Un blocco con molte voci
# dense puo' quindi restare ben sotto CHUNK_MAX_CHARS eppure produrre un
# output che tronca comunque (verificato dal vivo sul blocco 6 del
# documento #9: 7564 caratteri ma 21 voci tabellari dense, troncato anche
# dopo 3 dimezzamenti per caratteri -- il dimezzamento restava ancorato ai
# caratteri di input, mai al numero di voci da generare). CHUNK_MAX_ENTRIES
# limita quante righe non vuote entrano in un singolo blocco,
# indipendentemente da CHUNK_MAX_CHARS.
CHUNK_MAX_ENTRIES = 10
# Un 429 reale (a differenza del troncamento per lunghezza) puo' capitare anche
# con un pacing corretto, perche' la finestra di 60s e' condivisa a livello di
# organizzazione Groq (non solo dalle chiamate di questa estrazione) -- si
# riprova con un'attesa extra crescente invece di abortire subito l'estrazione.
CHUNK_RATE_LIMIT_MAX_RETRIES = 3
# erpv6.omni.provider.is_available() (omni_provider.py) ha un circuit breaker
# con cooldown FISSO di 5 minuti su QUALSIASI errore del provider, non solo un
# 429 -- verificato dal vivo sul documento #9: dopo un 429 reale, un retry a
# 130s e' arrivato PRIMA che il cooldown finisse, quindi il circuit breaker ha
# bloccato la chiamata ancora prima di provare Groq (l'errore diventa
# generico "Tutti i provider hanno fallito. Ultimo errore: None", non un 429),
# e senza riconoscerlo come rate limit l'estrazione abortiva comunque.
# L'attesa deve superare SEMPRE questo cooldown, non il pacing Groq normale.
CHUNK_RATE_LIMIT_WAIT_SECONDS = 310
# Ogni tot blocchi di primo livello completati, le voci estratte finora vengono
# salvate/commitate e mandate in validazione (vedi library_document.py,
# _commit_kb_extraction_batch) invece di aspettare la fine dell'intero
# documento -- un fallimento successivo non fa perdere il lavoro gia' fatto.
KB_BATCH_SIZE = 5
# Nome fisso (non tradotto: usato come chiave di ricerca in get_or_create, una
# traduzione diversa per lingua romperebbe il match) della categoria di
# transito per le voci KB appena estratte, prima della validazione 6 Giudici --
# vedi erpv6_production/models/validation_session.py per lo spostamento nella
# categoria reale suggerita dall'AI al momento dell'approvazione.
KB_STAGING_CATEGORY_NAME = "KB estratte — in attesa di validazione"

# Fase 1C.2 del knowledge graph erpv6 (vedi docs/PLAN_knowledge_graph_phase1.md,
# sezione 1C.1 "Predicati") -- vocabolario controllato per le triple che l'AI
# puo' estrarre insieme a ogni voce KB. Fino al 22/08/2026 era una lista
# Python hardcodata (ALLOWED_TRIPLE_SHAPES) -- ogni estensione richiedeva
# modifica codice/compila/promuovi/riavvia. Ora vive nel modello
# erpv6.kg.triple.shape (vedi kg_triple_shape.py, stesso modulo): ogni riga
# con is_active=True e' una tripla (subject_type, predicate, object_type)
# ammessa in estrazione, gestibile da Denis via la vista "Vocabolario Triple
# KG" senza toccare codice. Qualunque combinazione non attiva in tabella va
# scartata lato Python (vedi _call_and_parse_chunk/_clean_triples), mai
# creata al volo solo perche' il prompt lo suggerisce -- il prompt e la
# validazione leggono la STESSA fonte (get_allowed_shapes/_text, vedi sotto)
# cosi' non possono andare fuori sincrono tra loro. Le 14 shape che vivevano
# qui (11 layer bandi + 3 layer vendita/comunicazione) sono state migrate
# come righe is_active=True in data/kg_triple_shape_seed_data.xml -- stesso
# comportamento, nessuna regressione.
#
# Cache di richiesta: una singola query al modello per ogni chiamata a
# extract_kb_entries (vedi _load_triple_vocabulary), il risultato viaggia
# poi come parametro lungo tutta la catena di chiamate (mai una query per
# ogni singola tripla controllata).

# Attributi minimi del Socio (PLAN, sezione "Socio -- dettaglio"): SOLO quelli
# rilevanti per l'idoneita' a bandi, minimizzazione dati esplicita -- nessun
# altro dato personale del socio va accettato anche se l'AI lo restituisce.
SHAREHOLDER_ATTRIBUTE_KEYS = {'ownership_percentage', 'age_range', 'birth_year', 'gender', 'source'}
# Vocabolario chiuso FaseVendita (PLAN 1C.1-bis) -- normalizzato il
# 22/08/2026 dai valori grezzi reali visti su PL/PC/DM (le uniche 3 famiglie
# che hanno questo campo). Sinonimi esatti si fondono nel valore canonico;
# valori compositi che coprono piu' fasi reali producono PIU' triple
# APPLIES_IN_PHASE (una per fase atomica coperta), mai un nuovo valore-nodo
# composito -- decisione esplicita di Denis 22/08/2026. "Prime fasi" (DM-12,
# id 227) tenuto DISTINTO da "Prima call": il contenuto reale di quella voce
# ("Non usa mai il tuo nome in nessun messaggio" / "Relazione ancora formale
# O stile comunicativo generico") descrive un periodo generico iniziale
# della relazione via messaggistica, non la telefonata specifica "prima
# call" -- forzarlo su "Prima call" sarebbe un'inferenza non giustificata
# dal testo.
FASE_VENDITA_CANONICAL = [
    'Pre-call', 'Prima call', 'Seconda call', 'Durante call', 'Post-prima call',
    'Post-proposta', 'Post-follow-up', 'Prime fasi', 'Qualsiasi fase',
]
# Chiave: valore grezzo reale (confronto case-insensitive, spazi esterni
# ignorati) -> lista di 1+ fasi canoniche (>1 solo per i 2 valori compositi
# reali trovati). Valori grezzi NON presenti qui non vengono scartati (non e'
# detto siano un errore, potrebbe essere un sinonimo non ancora catalogato):
# passano con l'oggetto originale, ma con un log esplicito per revisione
# umana -- mai in silenzio, mai una fusione indovinata.
_FASE_VENDITA_MAP = {
    'qualsiasi': ['Qualsiasi fase'],
    'qualsiasi fase': ['Qualsiasi fase'],
    'pre-call': ['Pre-call'],
    'prima call': ['Prima call'],
    'seconda call': ['Seconda call'],
    'durante call': ['Durante call'],
    'post-prima call': ['Post-prima call'],
    'post-proposta': ['Post-proposta'],
    'post-follow-up': ['Post-follow-up'],
    'prime fasi': ['Prime fasi'],
    'prima e seconda call': ['Prima call', 'Seconda call'],
    'post-proposta / post-follow-up': ['Post-proposta', 'Post-follow-up'],
}
# Descrizione leggibile delle triple ammesse ora si genera a runtime da
# erpv6.kg.triple.shape (vedi _load_triple_vocabulary), non piu' da una
# lista statica qui -- cosi' prompt e validazione Python restano ancorati
# alla STESSA tabella, non solo alla stessa lista Python.
_FASE_VENDITA_VALUES_TEXT = ", ".join(f'"{v}"' for v in FASE_VENDITA_CANONICAL)

# Rilevatore "vocabolario insufficiente" (proposto da Denis il 22/08/2026 dopo
# aver scoperto stanotte, solo con una revisione manuale, che 109 voci
# consecutive producevano 0 triple perche' appartenevano a un intero layer di
# contenuto -- metodologia vendita/comunicazione -- mai coperto da
# ALLOWED_TRIPLE_SHAPES). Invece di scoprirlo di nuovo solo a mano, N
# estrazioni consecutive a zero triple scrivono un avviso reale (changelog_
# tecnico + to-do umano), poi il contatore si azzera -- mai piu' di un avviso
# per ogni streak, e MAI un blocco della pipeline (l'estrazione continua
# comunque, questo e' solo un segnale per revisione umana, nessuna estensione
# automatica del vocabolario).
TRIPLE_DISCOVERY_THRESHOLD = 3
_TRIPLE_DISCOVERY_COUNTER_KEY = 'erpv6_kb.consecutive_empty_triple_extractions'
_TRIPLE_DISCOVERY_SAMPLE_KEY = 'erpv6_kb.consecutive_empty_triple_extractions_sample'
_TRIPLE_DISCOVERY_PROPOSED_KEY = 'erpv6_kb.consecutive_empty_triple_extractions_proposed'

EXTRACTION_SYSTEM_PROMPT = """Sei un motore di estrazione dati per la Knowledge Base aziendale di V6 Impresa.
Ti viene fornito il contenuto grezzo di un documento (tabellare o testuale) caricato da un consulente.
Estrai le voci di conoscenza strutturate SENZA inventare contenuto che non è presente nel testo fornito.
Non fare allusioni, inferenze implicite o supposizioni: riporta solo ciò che è esplicitamente scritto nel testo, senza sottintendere, ipotizzare o suggerire informazioni non dichiarate.
Se una riga/sezione non contiene conoscenza utile (es. intestazioni vuote), ignorala.

Per ogni voce identificabile produci un oggetto JSON con questi campi:
- "name": titolo breve della voce
- "kb_type": uno tra questi valori esatti, scegli il più adatto al contenuto: {kb_types}
  Se il contenuto non è chiaramente riconducibile a nessuno di questi, usa "metodo_v6".
- "category": nome breve di categoria (es. "IVA agricola", "Comunicazione DISC")
- "content": il testo completo della voce, fedele all'originale, non riassunto se non necessario
- "triples": array di relazioni ESPLICITAMENTE presenti nel testo (vuoto se non ce ne sono, non forzarle).
  Ogni tripla è un oggetto con:
  - "subject_type": uno tra questi tipi di nodo: {node_types}
  - "subject": nome/etichetta del soggetto così come compare nel testo (es. "Azienda Rossi SRL"). Se subject_type è "VoceKB" usa sempre la stringa "(questa voce)"
  - "predicate": il predicato
  - "object_type": uno tra gli stessi tipi di nodo
  - "object": nome/etichetta/valore dell'oggetto così come compare nel testo
  - "attributes": oggetto opzionale, SOLO per predicate="HAS_SHAREHOLDER" e SOLO con questi campi se presenti esplicitamente nel testo: "ownership_percentage", "age_range" oppure "birth_year", "gender" (solo se il testo lo lega esplicitamente a un requisito di bando, mai altrimenti), "source" (valorizza "visura_camerale" solo se il testo lo indica come tale). Nessun altro dato personale del socio, mai.
  USA SOLO queste combinazioni esatte (subject_type -predicate-> object_type), non inventarne altre: {allowed_shapes}
  Se nel testo trovi una relazione reale che non rientra in nessuna di queste combinazioni, NON forzarla dentro "triples": descrivila invece come stringa libera in un campo separato "proposed_new_predicates" (array di stringhe, es. "Cliente-HA_CONSULENTE_ESTERNO->Persona: ..."), mai aggiunta a "triples".

Note sui tipi di nodo del layer Vendita/Comunicazione (metodologia commerciale/comunicazione, non bandi):
- "FaseVendita": per "object" usa SOLO uno di questi valori canonici: {fase_vendita_values}. Se il testo indica una fase composita che copre più fasi (es. "prima e seconda call"), produci una tripla APPLIES_IN_PHASE separata per ciascuna fase coperta, non un valore nuovo o una fase inventata.
- "FrameworkTeorico": vocabolario APERTO, usa la stringa esatta (nome teoria/autore) così come compare nel testo, es. "Cialdini — Influence", "NLP", "SPIN Selling". Non forzarla in uno dei valori di FaseVendita: sono due cose diverse.
- "ProfiloDISCTipo": SOLO uno tra "Dominante", "Influente", "Stabile", "Analitico", e SOLO se il testo lega esplicitamente per nome una voce a quel profilo specifico. Se il testo cita solo "DISC model"/"modello DISC" in modo generico senza nominare un profilo, usa BASED_ON -> FrameworkTeorico con object "DISC model", non TYPICAL_OF_PROFILE.

{extra_notes}

Rispondi SOLO con un array JSON valido (anche vuoto se non trovi nulla di utile), senza markdown code fence, senza altro testo."""


class Erpv6KbExtractionService(models.AbstractModel):
    """Motore condiviso: documento grezzo -> voci erpv6.kb strutturate via AI.

    Riusato sia dal wizard manuale (erpv6.kb.import.wizard) sia dal trigger
    automatico su erpv6.library.document (categoria 'kb_source') -- unica
    implementazione, nessuna duplicazione della logica di parsing/prompt/creazione.
    """
    _name = 'erpv6.kb.extraction.service'
    _description = 'Servizio di Estrazione KB da Documento (AI)'

    @api.model
    def extract_raw_text(self, file_data_b64, filename):
        if not file_data_b64:
            raise UserError(_("Nessun file fornito."))
        data = base64.b64decode(file_data_b64)
        filename = (filename or '').lower()

        if filename.endswith(('.xlsx', '.xlsm')):
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise UserError(_("Libreria openpyxl non disponibile sul server."))
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines = []
            for sheet in wb.worksheets:
                lines.append(f"--- Foglio: {sheet.title} ---")
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    cells = [str(c) for c in row if c is not None and str(c).strip()]
                    if cells:
                        lines.append(f"Riga {row_idx}: " + " | ".join(cells))
            return "\n".join(lines)

        if filename.endswith('.pdf'):
            try:
                from PyPDF2 import PdfReader
            except ImportError:
                raise UserError(_("Libreria PyPDF2 non disponibile sul server."))
            reader = PdfReader(io.BytesIO(data))
            pages = [(page.extract_text() or '') for page in reader.pages]
            # Riga vuota tra le pagine (non un singolo \n): e' il confine di
            # sezione reale piu' affidabile che abbiamo su un PDF senza
            # struttura esplicita, usato da _split_into_chunks per non
            # spezzare mai una pagina a meta' finche' non e' necessario.
            text = "\n\n".join(pages).strip()
            if not text:
                raise UserError(_(
                    "Il PDF non contiene testo estraibile (probabile scansione/immagine "
                    "senza OCR): questo server non ha un motore OCR configurato."
                ))
            return text

        if filename.endswith('.docx'):
            import zipfile
            from xml.etree import ElementTree
            try:
                with zipfile.ZipFile(io.BytesIO(data)) as z:
                    xml_bytes = z.read('word/document.xml')
            except (zipfile.BadZipFile, KeyError):
                raise UserError(_("File .docx non valido o corrotto."))
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            root = ElementTree.fromstring(xml_bytes)
            paragraphs = []
            for p in root.iter(f"{{{ns['w']}}}p"):
                texts = [t.text or '' for t in p.iter(f"{{{ns['w']}}}t")]
                paragraphs.append(''.join(texts))
            # Riga vuota tra i paragrafi (non un singolo \n): stesso motivo
            # del PDF sopra, e' il vero confine di paragrafo del .docx.
            return "\n\n".join(paragraphs)

        try:
            text = data.decode('utf-8')
        except UnicodeDecodeError:
            text = data.decode('latin-1', errors='replace')

        # Formati testuali generici (csv, xml, txt, md, ...) arrivano qui e
        # vanno bene cosi' come stringa grezza -- il prompt AI a valle decide
        # come interpretarli. Ma un file binario non riconosciuto sopra
        # (immagine, .doc legacy, ecc.) decodificato "a forza" con latin-1
        # produce sempre una stringa (non solleva mai UnicodeDecodeError),
        # quindi senza questo controllo verrebbe inviato all'AI come testo
        # illeggibile, sprecando la chiamata e producendo voci KB spazzatura.
        printable_ratio = sum(1 for c in text if c.isprintable() or c in '\n\r\t') / max(len(text), 1)
        if printable_ratio < 0.85:
            raise UserError(_(
                "Formato file non supportato per l'estrazione automatica del testo "
                "(probabile immagine o binario senza un parser dedicato). Formati "
                "supportati oggi: testo semplice, CSV, XML, .xlsx/.xlsm, .pdf (con testo, "
                "non scansioni), .docx."
            ))
        return text

    @api.model
    def _split_into_sections(self, text):
        """Divide il testo in sezioni "atomiche" che _split_into_chunks non
        spezzera' mai (a meno che una sezione da sola non superi
        max_chars) -- una sezione e' un blocco tra righe vuote (paragrafo
        .docx, pagina .pdf, vedi extract_raw_text) oppure un foglio .xlsx
        (marcatore "--- Foglio: " gia' presente nel testo, mai inventato).
        Testo senza alcuna riga vuota (es. .txt/.csv senza struttura)
        produce una sola sezione -- degrada a comportamento identico a prima
        (l'intera stringa passa dal fallback di taglio per lunghezza)."""
        # Split su 2+ a-capo consecutivi, MA forza anche un confine subito
        # prima di ogni "--- Foglio:" (l'export xlsx non mette righe vuote
        # tra le righe di uno stesso foglio, solo tra fogli diversi).
        text = re.sub(r'(?m)^(--- Foglio: )', r'\n\n\1', text)
        raw_sections = re.split(r'\n\s*\n+', text)
        return [s for s in (section.strip('\n') for section in raw_sections) if s]

    @api.model
    def _count_entries(self, text):
        """Righe non vuote in text -- proxy generico per il numero di
        "voci" da estrarre (per .xlsx corrisponde esattamente al numero di
        righe tabellari, una per voce KB, vedi extract_raw_text; per PDF/
        docx e' una stima piu' approssimativa ma comunque un argine utile
        contro blocchi troppo densi di output, vedi CHUNK_MAX_ENTRIES)."""
        return sum(1 for line in text.splitlines() if line.strip())

    @api.model
    def _limit_end_by_entries(self, text, start, end, max_lines):
        """Restringe [start:end) al piu' vicino a-capo dopo la
        max_lines-esima riga non vuota, se il segmento ne contiene di piu'
        -- non allarga mai end oltre il valore gia' calcolato per
        rispettare max_chars, lo restringe solo se necessario."""
        count = 0
        pos = start
        for line in text[start:end].splitlines(keepends=True):
            if line.strip():
                count += 1
            pos += len(line)
            if count >= max_lines:
                return pos
        return end

    @api.model
    def _split_oversized_section(self, text, max_chars, max_lines=None):
        """Spezza una singola sezione troppo grande per un blocco, su un
        a-capo vicino al limite quando possibile (stesso fallback di sempre,
        per non tagliare una riga di tabella a meta' e confondere il
        modello). Se max_lines e' passato, restringe ulteriormente ogni
        pezzo a al piu' max_lines righe non vuote (vedi CHUNK_MAX_ENTRIES)."""
        chunks = []
        start = 0
        length = len(text)
        while start < length:
            end = min(start + max_chars, length)
            if end < length:
                newline_pos = text.rfind('\n', start, end)
                if newline_pos > start:
                    end = newline_pos + 1
            if max_lines:
                end = self._limit_end_by_entries(text, start, end, max_lines)
            chunks.append(text[start:end])
            start = end
        return chunks

    @api.model
    def _split_into_chunks(self, text, max_chars, max_lines=None):
        """Pacchettizza il testo per sezioni intere (vedi
        _split_into_sections) invece che per taglio cieco a N caratteri: una
        sezione entra intera in un blocco finche' c'e' spazio, e viene
        spezzata internamente (_split_oversized_section, fallback identico
        al vecchio comportamento) solo se da sola supera max_chars O
        max_lines (voci non vuote, vedi CHUNK_MAX_ENTRIES). Cosi' i blocchi
        mandati ai 6 Giudici sono sezioni coerenti e complete, non frammenti
        a meta' capitolo, finche' non e' l'unica strada per rispettare il
        limite di output del modello -- ne' per caratteri (input) ne' per
        numero di voci (proxy dell'output, vedi CHUNK_MAX_ENTRIES)."""
        sections = self._split_into_sections(text)
        chunks = []
        current = ''
        for section in sections:
            oversized = len(section) > max_chars or (
                max_lines and self._count_entries(section) > max_lines)
            if oversized:
                if current:
                    chunks.append(current)
                    current = ''
                chunks.extend(self._split_oversized_section(section, max_chars, max_lines))
                continue
            candidate = f"{current}\n\n{section}" if current else section
            candidate_oversized = len(candidate) > max_chars or (
                max_lines and self._count_entries(candidate) > max_lines)
            if candidate_oversized and current:
                chunks.append(current)
                current = section
            else:
                current = candidate
        if current:
            chunks.append(current)
        return chunks

    @api.model
    def estimate_extraction_plan(self, file_data_b64, filename):
        """Stima blocchi/tempo attesi PRIMA di lanciare l'estrazione vera e
        propria (sincrona, vedi extract_kb_entries) -- usata dal chiamante
        (erpv6_production/models/library_document.py) per notificare
        subito l'utente di quanto durera' probabilmente, invece di lasciarlo
        senza segnali per diversi minuti su un documento grande. Riusa
        extract_raw_text/_split_into_chunks (stessa logica di
        extract_kb_entries, nessuna duplicazione) -- solleva la stessa
        UserError di extract_raw_text se il file non e' leggibile; il
        chiamante puo' ignorarla e lasciare che extract_kb_entries la
        risollevi (parsing di un file e' economico, non ripete alcuna
        chiamata AI)."""
        raw_text = self.extract_raw_text(file_data_b64, filename)
        chunk_count = len(self._split_into_chunks(raw_text, CHUNK_MAX_CHARS, max_lines=CHUNK_MAX_ENTRIES))
        return {
            'chunk_count': chunk_count,
            'estimated_seconds': chunk_count * CHUNK_PACING_SECONDS,
        }

    @api.model
    def _resolve_route_and_model(self):
        # Il payload viene inviato cosi' com'e' al provider che il bridge sceglie
        # (vedi erpv6.omni.bridge.execute_ai_task, che non adatta 'model' in base
        # al provider effettivo) -- usiamo il modello del provider primario di
        # QUESTA route invece di un default fisso, cosi' almeno il percorso
        # primario ha un nome di modello valido. Se il primario fallisce e il
        # bridge passa a un fallback, quel provider potrebbe non avere questo
        # stesso modello -- gap noto nel bridge, non risolvibile da qui senza
        # cambiare execute_ai_task per ogni altra route che lo usa.
        route = self.env['erpv6.omni.route.config'].search([
            ('task_type', '=', 'kb_extraction'), ('is_active', '=', True),
        ], limit=1)
        model = route.primary_provider_id.get_optimal_model('kb_extraction') if route and route.primary_provider_id else None
        return model or 'gpt-4-turbo'

    @api.model
    def _call_and_parse_chunk(self, system_prompt, chunk_text, model, chunk_label, allowed_shapes, max_tokens=None):
        payload = {
            'model': model,
            'temperature': 0.1,
            'max_tokens': max_tokens or CHUNK_MAX_OUTPUT_TOKENS,
            'messages': [
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': chunk_text},
            ],
        }

        bridge = self.env['erpv6.omni.bridge']
        result = bridge.execute_ai_task(
            task_type='kb_extraction',
            payload=payload,
            context={'source': 'kb_extraction_service'},
        )

        if not result.get('success'):
            error_text = result.get('error', 'errore sconosciuto')
            # execute_ai_task propaga str(HTTPError) da requests.raise_for_status()
            # cosi' com'e' -- un 429 reale contiene sempre questo testo, distinto
            # da qualsiasi altro errore (chiave mancante, formato, timeout...).
            # "Ultimo errore: None" e' il secondo segnale: il circuit breaker di
            # is_available() ha bloccato TUTTI i provider prima ancora di
            # provarli (cooldown post-429 non ancora scaduto, vedi
            # CHUNK_RATE_LIMIT_WAIT_SECONDS) -- anche questo e' transitorio,
            # NON un errore di configurazione reale (quello avrebbe un
            # last_error concreto, non None, dopo "Ultimo errore:").
            circuit_breaker_blocked = error_text.rstrip().endswith('Ultimo errore: None')
            if '429' in error_text or 'too many requests' in error_text.lower() or circuit_breaker_blocked:
                raise _ChunkRateLimitedError(f"{chunk_label}: {error_text}")
            raise UserError(_(
                "Estrazione AI fallita (%(chunk)s): %(error)s\n\n"
                "Nota: se nessun provider AI (OpenAI/Anthropic/Groq) ha una API Key attiva "
                "configurata in Impostazioni > V6 Impresa AI > OmniRoute > Provider, questo "
                "errore è atteso finché non ne viene attivato almeno uno."
            ) % {'chunk': chunk_label, 'error': error_text})

        raw_content = result.get('data', {})
        try:
            choice = raw_content['choices'][0]
            message_content = choice['message']['content']
        except (KeyError, IndexError, TypeError):
            raise UserError(_(
                "Risposta AI in un formato inatteso (%(chunk)s), non riconducibile a una chat "
                "completion standard. Risposta grezza: %(raw)s"
            ) % {'chunk': chunk_label, 'raw': json.dumps(raw_content)[:1000]})

        # finish_reason='length' (schema OpenAI-compatibile, valido anche per
        # Groq) e' l'unico segnale affidabile che l'output e' stato tagliato
        # per CHUNK_MAX_OUTPUT_TOKENS -- serve a distinguere questo caso
        # (riprovabile dimezzando il blocco) da un JSON davvero malformato
        # (finish_reason='stop', errore reale non riprovabile allo stesso modo).
        truncated_by_length = choice.get('finish_reason') == 'length'

        if message_content is None:
            # Modelli "reasoning" (es. Gemini 3.6) possono restituire
            # message.content=None con finish_reason='length' quando il
            # budget di max_tokens si esaurisce nel ragionamento interno
            # prima di produrre testo -- stesso rimedio del troncamento
            # normale (dimezzare il blocco riduce anche il ragionamento
            # necessario). Se invece finish_reason non e' 'length', un
            # content nullo e' davvero inatteso: mai inventare un
            # risultato vuoto, va segnalato con la risposta grezza.
            if truncated_by_length:
                raise _ChunkTruncatedError(
                    f"{chunk_label}: nessun contenuto testuale nella risposta "
                    "(probabile budget di reasoning esaurito) con finish_reason=length"
                )
            raise UserError(_(
                "Risposta AI senza contenuto testuale (%(chunk)s), finish_reason=%(reason)s. "
                "Risposta grezza: %(raw)s"
            ) % {'chunk': chunk_label, 'reason': choice.get('finish_reason'), 'raw': json.dumps(raw_content)[:1000]})

        message_content = message_content.strip()
        if message_content.startswith('```'):
            message_content = message_content.strip('`')
            if message_content.lower().startswith('json'):
                message_content = message_content[4:]
            message_content = message_content.strip()

        try:
            entries = json.loads(message_content)
        except json.JSONDecodeError as e:
            if truncated_by_length:
                raise _ChunkTruncatedError(
                    f"{chunk_label}: output troncato per limite lunghezza (finish_reason=length)")
            # message_content[:1000] (comportamento precedente) mostra SEMPRE
            # solo l'inizio della risposta -- inutile per un errore tipo
            # "Extra data" (es. un ']' di troppo in coda), la cui posizione
            # (e.pos) cade quasi sempre ben oltre i primi 1000 caratteri su
            # una risposta reale con voci KB + triple (verificato dal vivo il
            # 22/08/2026: una risposta di ~4600 caratteri fallisce a pos
            # ~4665, quindi il vecchio troncamento nascondeva sempre proprio
            # la parte utile a diagnosticare il problema). Logga qui il
            # contenuto grezzo COMPLETO (mai troncato) lato server per la
            # prossima occorrenza, e mostra nel messaggio utente il contesto
            # attorno alla posizione reale dell'errore invece dell'inizio
            # fisso della stringa.
            _logger.error(
                "Estrazione KB (%s): JSON non valido (%s), contenuto grezzo COMPLETO per diagnosi:\n%s",
                chunk_label, e, message_content)
            error_context = message_content[max(0, e.pos - 200):e.pos + 200]
            raise UserError(_(
                "L'AI non ha risposto con un JSON valido (%(chunk)s): %(err)s\n\n"
                "Contesto attorno alla posizione dell'errore (car. %(pos)d di %(total)d): %(context)s\n\n"
                "Contenuto grezzo completo salvato nei log del server per diagnosi."
            ) % {
                'chunk': chunk_label, 'err': e, 'pos': e.pos, 'total': len(message_content),
                'context': error_context,
            })

        if not isinstance(entries, list):
            raise UserError(_("L'AI non ha restituito un array di voci (%s).") % chunk_label)

        valid_kb_types = {k for k, _label in KB_TYPE_SELECTION}
        cleaned = []
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            kb_type = entry.get('kb_type') if entry.get('kb_type') in valid_kb_types else 'metodo_v6'
            triples = self._clean_triples(entry.get('triples'), chunk_label, entry.get('name'), allowed_shapes)
            cleaned.append({
                'name': entry.get('name') or _('Senza titolo'),
                'kb_type': kb_type,
                'category': entry.get('category') or _('Generale'),
                'content': entry.get('content') or '',
                'triples': triples,
            })
            proposed_new_predicates = entry.get('proposed_new_predicates')
            self._log_proposed_triple_extensions(proposed_new_predicates, chunk_label, entry.get('name'))
            self._track_triple_discovery(triples, entry.get('name'), kb_type, proposed_new_predicates)
        return cleaned

    @api.model
    def _track_triple_discovery(self, triples, entry_name, kb_type, proposed_new_predicates=None):
        """Contatore persistente (ir.config_parameter, sopravvive tra
        chiamate/documenti diversi) di estrazioni CONSECUTIVE senza nessuna
        tripla in ambito. Ogni voce con almeno 1 tripla resetta il contatore
        a 0 (il vocabolario ha funzionato, non serve un avviso). Al
        raggiungimento di TRIPLE_DISCOVERY_THRESHOLD, scrive un avviso reale
        e si riazzera -- mai un secondo avviso identico per la stessa streak.
        Accumula anche (stesso meccanismo persistente) i valori grezzi di
        proposed_new_predicates visti nelle voci a zero triple della streak
        corrente -- passati a _raise_triple_discovery_signal, che puo' usarli
        per dedurre una proposta di shape concreta (vedi sotto)."""
        Param = self.env['ir.config_parameter'].sudo()
        if triples:
            Param.set_param(_TRIPLE_DISCOVERY_COUNTER_KEY, '0')
            Param.set_param(_TRIPLE_DISCOVERY_SAMPLE_KEY, '')
            Param.set_param(_TRIPLE_DISCOVERY_PROPOSED_KEY, '')
            return
        count = int(Param.get_param(_TRIPLE_DISCOVERY_COUNTER_KEY, '0') or '0')
        sample_raw = Param.get_param(_TRIPLE_DISCOVERY_SAMPLE_KEY, '') or ''
        sample = [s for s in sample_raw.split('||') if s] if sample_raw else []
        sample.append(f"{entry_name or '(senza nome)'} [{kb_type}]")
        proposed_raw = Param.get_param(_TRIPLE_DISCOVERY_PROPOSED_KEY, '') or ''
        proposed_accum = [s for s in proposed_raw.split('||') if s] if proposed_raw else []
        if isinstance(proposed_new_predicates, list):
            proposed_accum.extend(
                item.strip() for item in proposed_new_predicates
                if isinstance(item, str) and item.strip())
        count += 1
        if count >= TRIPLE_DISCOVERY_THRESHOLD:
            self._raise_triple_discovery_signal(sample, proposed_accum)
            Param.set_param(_TRIPLE_DISCOVERY_COUNTER_KEY, '0')
            Param.set_param(_TRIPLE_DISCOVERY_SAMPLE_KEY, '')
            Param.set_param(_TRIPLE_DISCOVERY_PROPOSED_KEY, '')
        else:
            Param.set_param(_TRIPLE_DISCOVERY_COUNTER_KEY, str(count))
            Param.set_param(_TRIPLE_DISCOVERY_SAMPLE_KEY, '||'.join(sample))
            Param.set_param(_TRIPLE_DISCOVERY_PROPOSED_KEY, '||'.join(proposed_accum))

    # Formato suggerito nel prompt per "proposed_new_predicates" (vedi
    # EXTRACTION_SYSTEM_PROMPT: 'es. "Cliente-HA_CONSULENTE_ESTERNO->Persona: ..."')
    # -- riconosce SOLO questo pattern esplicito "Soggetto-PREDICATO->Oggetto"
    # (descrizione libera opzionale dopo ":"), mai un'inferenza piu' furba sul
    # testo libero: se la stringa non rispetta il formato, niente proposta
    # automatica (vedi _parse_proposed_predicate_shape).
    _PROPOSED_PREDICATE_RE = re.compile(r'^([A-Za-z]\w*)-([A-Z][A-Z0-9_]*)->([A-Za-z]\w*)')

    @api.model
    def _parse_proposed_predicate_shape(self, raw_text):
        """Tenta di dedurre (subject_type, predicate, object_type) da una
        singola stringa proposed_new_predicates. Ritorna None se il testo
        non rispetta esplicitamente il formato "Soggetto-PREDICATO->Oggetto"
        -- MAI un'inferenza indovinata: senza questo pattern chiaro, nessuna
        riga proposta viene creata per quella stringa (resta comunque nel
        changelog testuale/log, vedi _log_proposed_triple_extensions)."""
        if not isinstance(raw_text, str):
            return None
        match = self._PROPOSED_PREDICATE_RE.match(raw_text.strip())
        if not match:
            return None
        return match.group(1), match.group(2), match.group(3)

    @api.model
    def _raise_triple_discovery_signal(self, sample_entries, proposed_predicates=None):
        """Scrive un avviso reale (voce erpv6.kb changelog_tecnico + to-do
        umano) quando il vocabolario controllato non ha trovato nessuna
        tripla per TRIPLE_DISCOVERY_THRESHOLD voci consecutive -- segnale che
        potrebbe servire un nuovo layer di vocabolario (vedi l'analisi del
        22/08/2026 sul layer "vendita/comunicazione", trovato con questo
        stesso ragionamento ma a mano).

        In aggiunta all'avviso testuale (sempre scritto), tenta di dedurre
        da proposed_predicates (i valori grezzi di "proposed_new_predicates"
        visti nella streak corrente, vedi _track_triple_discovery) una o piu'
        shape concrete via _parse_proposed_predicate_shape, e le scrive come
        righe PROPOSTA (is_active=False, source='discovery_proposal') in
        erpv6.kg.triple.shape tramite propose_discovery_shape -- MAI attive:
        restano sempre un click umano (vedi action_activate sul modello).
        Se non riesce a dedurre nulla di specifico da nessuna stringa, resta
        comunque solo l'avviso testuale, come prima di questa estensione."""
        category = self.env['erpv6.kb.category'].get_or_create(
            "Changelog Tecnico", 'changelog_tecnico', is_transversal=True)

        shape_model = self.env['erpv6.kg.triple.shape']
        created_proposals = []
        seen_shapes = set()
        for raw in (proposed_predicates or []):
            parsed = self._parse_proposed_predicate_shape(raw)
            if not parsed or parsed in seen_shapes:
                continue
            seen_shapes.add(parsed)
            subject_type, predicate, object_type = parsed
            proposal = shape_model.propose_discovery_shape(
                subject_type, predicate, object_type,
                notes=_("Dedotta automaticamente dal rilevatore vocabolario insufficiente "
                        "da proposed_new_predicates: %s") % raw,
            )
            if proposal:
                created_proposals.append(proposal)
                _logger.info(
                    "Estrazione KB: proposta di shape creata/riusata da rilevatore automatico: "
                    "%s -%s-> %s (id=%s, is_active=%s)",
                    subject_type, predicate, object_type, proposal.id, proposal.is_active)

        proposals_text = ""
        if created_proposals:
            proposals_text = _(
                "\n\nDedotte %(n)d proposte di shape (inattive, da rivedere e attivare a mano "
                "in AI & OmniRoute > Configurazione > Vocabolario Triple KG): %(list)s"
            ) % {
                'n': len(created_proposals),
                'list': '; '.join(p.display_name for p in created_proposals),
            }

        entry = self.env['erpv6.kb'].create({
            'name': _("Possibile vocabolario KG insufficiente (%s voci consecutive senza triple)") % TRIPLE_DISCOVERY_THRESHOLD,
            'kb_type': 'changelog_tecnico',
            'category_id': category.id,
            'content_format': 'text',
            'source': 'kb_extraction_service:_raise_triple_discovery_signal',
            'is_active': True,
            'content': _(
                "%(n)s voci KB consecutive estratte con zero triple in ambito (nessuna "
                "combinazione soggetto/predicato/oggetto attiva nel vocabolario controllato "
                "erpv6.kg.triple.shape). Non è necessariamente un problema — può essere "
                "contenuto genuinamente fuori da qualunque vocabolario — ma vale la pena "
                "verificare se serve un nuovo layer di vocabolario (vedi PLAN sezione 1C.1, e "
                "l'analisi del 22/08/2026 sul layer 'vendita/comunicazione' trovato con questo "
                "stesso metodo, a mano). Ultime voci coinvolte: %(sample)s%(proposals)s"
            ) % {
                'n': TRIPLE_DISCOVERY_THRESHOLD, 'sample': '; '.join(sample_entries),
                'proposals': proposals_text,
            },
        })
        # Notifica tramite l'agente Sabrina (erpv6_agent, code='sabrina') invece
        # di un avviso generico di sistema -- richiesto esplicitamente da Denis
        # il 22/08/2026 ("voglio che sia Sabrina che mi scrive"). Se il rilevatore
        # e' riuscito a dedurre almeno una proposta concreta, ognuna diventa una
        # erpv6.agent.confirmation separata con l'azione predefinita gia'
        # collegata (action_activate su quella riga) -- Denis conferma con una
        # parola chiave nello stesso thread e SOLO quella riga si attiva, mai
        # un'interpretazione libera. Se non c'e' nessuna proposta concreta, resta
        # comunque un avviso a voce di Sabrina, ma senza azione da confermare
        # (confirmed_no_action se risposto, vedi erpv6.agent.confirmation).
        sabrina = self.env['erpv6.agent.config'].sudo().search([('code', '=', 'sabrina')], limit=1)
        facts = [
            _("%(n)s voci KB consecutive senza nessuna tripla trovata dal motore di estrazione.")
            % {'n': TRIPLE_DISCOVERY_THRESHOLD},
            _("Potrebbe indicare un intero tipo di contenuto senza vocabolario dedicato "
              "(come successo stanotte con la metodologia vendita)."),
            _("Ultime voci coinvolte: %s") % '; '.join(sample_entries),
        ]
        if sabrina:
            if created_proposals:
                for proposal in created_proposals:
                    sabrina.notify_pending_confirmation(
                        title=_("Possibile vocabolario KG insufficiente — proposta: %s") % proposal.display_name,
                        facts=facts + [_(
                            "Proposta dedotta: %s. Rispondi con una parola chiave (es. \"ok\") "
                            "in questo stesso messaggio per attivarla, oppure ignora/rispondi "
                            "diversamente per lasciarla in sospeso."
                        ) % proposal.display_name],
                        res_model='erpv6.kb', res_id=entry.id,
                        action_model='erpv6.kg.triple.shape', action_res_id=proposal.id,
                        action_method='action_activate',
                    )
            else:
                sabrina.notify_pending_confirmation(
                    title=_("Possibile vocabolario KG insufficiente"),
                    facts=facts + [_(
                        "Nessuna proposta concreta dedotta automaticamente — rivedi le voci "
                        "coinvolte a mano e valuta se serve un nuovo layer di vocabolario."
                    )],
                    res_model='erpv6.kb', res_id=entry.id,
                )
        else:
            _logger.warning(
                "Estrazione KB: agente Sabrina non trovato (erpv6.agent.config code='sabrina'), "
                "avviso vocabolario insufficiente scritto solo nel changelog KB #%s, nessuna "
                "notifica inviata.", entry.id)
        _logger.info(
            "Estrazione KB: rilevatore vocabolario insufficiente attivato dopo %s voci "
            "consecutive senza triple, voce changelog #%s creata, %d proposte di shape dedotte.",
            TRIPLE_DISCOVERY_THRESHOLD, entry.id, len(created_proposals))

    @api.model
    def _clean_triples(self, raw_triples, chunk_label, entry_name, allowed_shapes):
        """Filtra le triple proposte dall'AI per una voce sul vocabolario
        controllato (Fase 1C.2, vedi docs/PLAN_knowledge_graph_phase1.md),
        letto dal chiamante da erpv6.kg.triple.shape (vedi
        _load_triple_vocabulary/extract_kb_entries) e passato qui come
        insieme di tuple -- MAI fidarsi solo del prompt: qualunque
        combinazione (subject_type, predicate, object_type) non presente in
        allowed_shapes viene scartata qui, con un log per tracciabilita'
        (non un errore bloccante: una tripla rifiutata non deve far fallire
        l'intera estrazione della voce, che resta comunque valida).

        :param allowed_shapes: set/lista di tuple (subject_type, predicate,
            object_type) attive -- calcolata UNA volta per estrazione dal
            chiamante, mai una query per singola tripla qui."""
        if not isinstance(raw_triples, list):
            return []
        cleaned_triples = []
        for triple in raw_triples:
            if not isinstance(triple, dict):
                continue
            subject_type = triple.get('subject_type')
            predicate = triple.get('predicate')
            object_type = triple.get('object_type')
            subject = triple.get('subject')
            obj = triple.get('object')
            if not all(isinstance(v, str) and v.strip() for v in (subject_type, predicate, object_type, subject, obj)):
                _logger.warning(
                    "Estrazione KB (%s, voce '%s'): tripla scartata, campi mancanti/non testuali: %r",
                    chunk_label, entry_name, triple)
                continue
            if (subject_type, predicate, object_type) not in allowed_shapes:
                _logger.warning(
                    "Estrazione KB (%s, voce '%s'): tripla scartata, combinazione fuori vocabolario "
                    "controllato: %s -%s-> %s (mai creata al volo)",
                    chunk_label, entry_name, subject_type, predicate, object_type)
                continue
            cleaned = {
                'subject_type': subject_type, 'subject': subject.strip(),
                'predicate': predicate, 'object_type': object_type, 'object': obj.strip(),
            }
            if predicate == 'HAS_SHAREHOLDER' and isinstance(triple.get('attributes'), dict):
                # Minimizzazione dati esplicita (PLAN, "Socio -- dettaglio"):
                # solo le chiavi whitelisted, tutto il resto scartato anche
                # se l'AI lo restituisce.
                attrs = {k: v for k, v in triple['attributes'].items()
                          if k in SHAREHOLDER_ATTRIBUTE_KEYS and isinstance(v, (str, int, float))}
                if attrs:
                    cleaned['attributes'] = attrs
            if predicate == 'APPLIES_IN_PHASE':
                # Normalizzazione FaseVendita (PLAN 1C.1-bis, decisione
                # esplicita di Denis 22/08/2026): non fidarsi del valore
                # letterale dell'AI, normalizzare qui lato Python -- stessa
                # disciplina gia' applicata sopra per HAS_SHAREHOLDER. Un
                # valore composito reale (es. "Prima e seconda call")
                # produce PIU' triple, una per fase atomica coperta, mai un
                # nodo-fase composito nuovo.
                cleaned_triples.extend(
                    self._normalize_fase_vendita_triple(cleaned, chunk_label, entry_name))
                continue
            cleaned_triples.append(cleaned)
        return cleaned_triples

    @api.model
    def _normalize_fase_vendita_triple(self, cleaned, chunk_label, entry_name=None):
        """Espande/normalizza una singola tripla APPLIES_IN_PHASE gia'
        validata sulla FORMA (shape gia' controllata dal chiamante) usando
        _FASE_VENDITA_MAP per il VALORE dell'oggetto. Ritorna una lista (1
        elemento normale, 2 per i valori compositi reali, 1 con l'oggetto
        originale intatto se il valore grezzo non e' nella mappa -- mai
        scartata in silenzio, mai fusa per indovinare: solo loggata per
        revisione umana, cosi' la mappa puo' crescere con dati reali)."""
        raw_key = cleaned['object'].strip().lower()
        canonical_values = _FASE_VENDITA_MAP.get(raw_key)
        if canonical_values is None:
            _logger.warning(
                "Estrazione KB (%s, voce '%s'): valore FaseVendita non nel vocabolario "
                "canonico normalizzato, tenuto cosi' com'e' per revisione umana: %r",
                chunk_label, entry_name, cleaned['object'])
            return [cleaned]
        return [dict(cleaned, object=value) for value in canonical_values]

    @api.model
    def _log_proposed_triple_extensions(self, proposed, chunk_label, entry_name=None):
        """Le proposte di predicati/tipi-nodo fuori dal vocabolario controllato
        (campo "proposed_new_predicates" del prompt, vedi EXTRACTION_SYSTEM_PROMPT)
        non vengono MAI create al volo ne' persistite QUI -- questo metodo si
        limita a loggarle per visibilita' durante il campione reale. La sola
        strada per farle diventare una riga (sempre inattiva, sempre da
        attivare a mano) in erpv6.kg.triple.shape e' via il rilevatore
        _track_triple_discovery/_raise_triple_discovery_signal, quando la
        stessa stringa ricompare in una streak di voci senza triple."""
        if not proposed:
            return
        if not isinstance(proposed, list):
            return
        for item in proposed:
            if isinstance(item, str) and item.strip():
                _logger.info(
                    "Estrazione KB (%s, voce '%s'): predicato/tipo fuori vocabolario proposto "
                    "dall'AI (NON creato, solo segnalato): %s",
                    chunk_label, entry_name, item.strip())

    @api.model
    def _extract_chunk_with_split_retry(self, system_prompt, chunk_text, model, chunk_label, call_state,
                                          allowed_shapes, depth=0, max_tokens=None):
        """Chiama _call_and_parse_chunk con pacing condiviso (call_state['calls']
        conta le chiamate AI reali fatte finora in questa estrazione, per
        rispettare CHUNK_PACING_SECONDS anche tra le sotto-chiamate di uno split,
        non solo tra i blocchi di primo livello). Gestisce due errori riprovabili
        in modo diverso: su troncamento per lunghezza (_ChunkTruncatedError)
        dimezza il blocco e riprova ricorsivamente fino a CHUNK_SPLIT_MAX_DEPTH
        (un blocco denso si adatta da solo); su rate limit reale
        (_ChunkRateLimitedError) NON dimezza (il blocco non e' il problema) ma
        riprova lo stesso blocco con un'attesa extra crescente, fino a
        CHUNK_RATE_LIMIT_MAX_RETRIES."""
        if call_state['calls'] > 0:
            _logger.info(
                "Estrazione KB: pausa di %ss prima di %s (limite gratuito Groq)",
                CHUNK_PACING_SECONDS, chunk_label)
            time.sleep(CHUNK_PACING_SECONDS)
        call_state['calls'] += 1
        effective_max_tokens = max_tokens or CHUNK_MAX_OUTPUT_TOKENS

        rate_limit_retries = 0
        while True:
            try:
                return self._call_and_parse_chunk(
                    system_prompt, chunk_text, model, chunk_label, allowed_shapes, effective_max_tokens)
            except _ChunkRateLimitedError as e:
                rate_limit_retries += 1
                if rate_limit_retries > CHUNK_RATE_LIMIT_MAX_RETRIES:
                    raise UserError(_(
                        "Limite di richieste Groq superato per %(chunk)s anche dopo %(n)d tentativi "
                        "extra di attesa: il servizio AI è sovraccarico, verosimilmente da altro "
                        "traffico sull'organizzazione Groq e non solo da questa estrazione. "
                        "Dettaglio: %(error)s"
                    ) % {'chunk': chunk_label, 'n': CHUNK_RATE_LIMIT_MAX_RETRIES, 'error': str(e)})
                # Deve sempre superare CHUNK_RATE_LIMIT_WAIT_SECONDS (il
                # cooldown fisso del circuit breaker, vedi sopra) -- un
                # pacing crescente ma piu' corto di quello e' garantito che
                # fallisca di nuovo, verificato dal vivo sul documento #9.
                wait_seconds = CHUNK_RATE_LIMIT_WAIT_SECONDS + CHUNK_PACING_SECONDS * (rate_limit_retries - 1)
                _logger.warning(
                    "Estrazione KB: rate limit/circuit breaker su %s, attesa extra di %ss (tentativo %d/%d)",
                    chunk_label, wait_seconds, rate_limit_retries, CHUNK_RATE_LIMIT_MAX_RETRIES)
                time.sleep(wait_seconds)
            except _ChunkTruncatedError:
                if depth >= CHUNK_SPLIT_MAX_DEPTH or len(chunk_text) < 500:
                    raise UserError(_(
                        "L'AI continua a troncare l'output per %(chunk)s anche dopo %(depth)d "
                        "tentativi di dimezzamento del blocco: il contenuto in questa porzione "
                        "del documento è troppo denso per il limite di output configurato. "
                        "Valuta di caricare questa parte del documento separatamente, in un file "
                        "più piccolo."
                    ) % {'chunk': chunk_label, 'depth': depth})
                half = max(len(chunk_text) // 2, 1)
                half_entries = max(self._count_entries(chunk_text) // 2, 1)
                sub_chunks = self._split_into_chunks(chunk_text, half, max_lines=half_entries)
                # Un sotto-blocco piu' piccolo non ha bisogno del budget di
                # output pieno riservato al blocco originale -- ridurlo
                # proporzionalmente abbassa anche il totale di token
                # richiesti (input+max_tokens) nella finestra scorrevole di
                # Groq, che e' esattamente cio' che ha fatto scattare il 429
                # reale sul documento #9 quando piu' blocchi si spezzavano
                # nella stessa finestra.
                sub_max_tokens = max(
                    CHUNK_MIN_OUTPUT_TOKENS,
                    min(CHUNK_MAX_OUTPUT_TOKENS, round(CHUNK_MAX_OUTPUT_TOKENS * half / CHUNK_MAX_CHARS)))
                _logger.info(
                    "Estrazione KB: %s troncato per lunghezza, diviso in %d sotto-blocchi "
                    "(tentativo %d/%d, max_tokens %d)",
                    chunk_label, len(sub_chunks), depth + 1, CHUNK_SPLIT_MAX_DEPTH, sub_max_tokens)
                entries = []
                for sub_idx, sub_text in enumerate(sub_chunks, start=1):
                    sub_label = f"{chunk_label}.{sub_idx}"
                    entries.extend(self._extract_chunk_with_split_retry(
                        system_prompt, sub_text, model, sub_label, call_state, allowed_shapes,
                        depth=depth + 1, max_tokens=sub_max_tokens))
                return entries

    @api.model
    def extract_kb_entries(self, file_data_b64, filename, notes=None, start_chunk_index=0, on_batch_complete=None):
        """Ritorna una lista di dict {name, kb_type, category, content} estratti via AI.

        Documenti piu' grandi di CHUNK_MAX_CHARS vengono spezzati in piu' chiamate
        sequenziali (col limite gratuito Groq -- vedi commento su CHUNK_PACING_SECONDS
        in testa al file) e i risultati aggregati. Solleva UserError se un qualsiasi
        chunk fallisce dopo aver esaurito i retry (nessun provider attivo, risposta
        non valida, rate limit persistente, ecc.) -- il chiamante decide come
        gestirlo (mostrare all'utente nel wizard, o loggarlo nel chatter del
        documento).

        :param start_chunk_index: indice (0-based) del primo blocco di primo
            livello da elaborare -- permette di riprendere un'estrazione fallita
            a meta' senza rielaborare (e duplicare) i blocchi gia' completati in
            un tentativo precedente (vedi library_document.py, che passa qui il
            campo persistito kb_extraction_completed_chunks).
        :param on_batch_complete: callback opzionale invocata ogni KB_BATCH_SIZE
            blocchi di primo livello completati (e sull'ultimo blocco del
            documento), con (entries_del_batch, indice_1based_ultimo_blocco).
            Permette al chiamante di salvare/commitare il lavoro incrementalmente
            invece di perdere tutto un documento lungo per un fallimento verso la
            fine (vedi library_document.py/_commit_kb_extraction_batch). Se non
            passata, il comportamento resta tutto-o-niente come prima.
        """
        raw_text = self.extract_raw_text(file_data_b64, filename)
        if not raw_text.strip():
            raise UserError(_("Nessun contenuto leggibile trovato nel documento."))

        # Vocabolario controllato letto UNA sola volta per questa estrazione
        # da erpv6.kg.triple.shape (vedi commento in testa al file) -- non
        # una query per ogni blocco/tripla, il risultato viaggia come
        # parametro lungo tutta la catena di chiamate sotto (_extract_chunk_
        # with_split_retry -> _call_and_parse_chunk -> _clean_triples).
        shape_model = self.env['erpv6.kg.triple.shape']
        allowed_shapes = shape_model.get_allowed_shapes()
        allowed_shapes_set = set(allowed_shapes)
        allowed_node_types = {t for shape in allowed_shapes for t in (shape[0], shape[2])}
        allowed_shapes_text = "; ".join(f"{s} -{p}-> {o}" for s, p, o in allowed_shapes)

        extra_notes = f"Note aggiuntive dell'utente: {notes}" if notes else ""
        system_prompt = EXTRACTION_SYSTEM_PROMPT.format(
            kb_types=", ".join(f'"{k}"' for k, _label in KB_TYPE_SELECTION),
            node_types=", ".join(f'"{t}"' for t in sorted(allowed_node_types)),
            allowed_shapes=allowed_shapes_text,
            fase_vendita_values=_FASE_VENDITA_VALUES_TEXT,
            extra_notes=extra_notes,
        )
        model = self._resolve_route_and_model()
        chunks = self._split_into_chunks(raw_text, CHUNK_MAX_CHARS, max_lines=CHUNK_MAX_ENTRIES)
        total = len(chunks)

        all_entries = []
        batch_entries = []
        call_state = {'calls': 0}
        for idx in range(start_chunk_index, total):
            chunk_number = idx + 1
            chunk_label = f"blocco {chunk_number}/{total}"
            entries = self._extract_chunk_with_split_retry(
                system_prompt, chunks[idx], model, chunk_label, call_state, allowed_shapes_set)
            all_entries.extend(entries)
            batch_entries.extend(entries)
            is_last_chunk = chunk_number == total
            if on_batch_complete and (chunk_number % KB_BATCH_SIZE == 0 or is_last_chunk):
                on_batch_complete(batch_entries, chunk_number)
                batch_entries = []

        return all_entries

    @api.model
    def create_kb_records(self, entries, source_label=None):
        """Crea (o riusa) le erpv6.kb.category necessarie e crea le erpv6.kb.

        Le voci vengono create nella categoria di transito KB_STAGING_CATEGORY_NAME
        (una per kb_type), NON nella categoria suggerita dall'AI (entry['category'],
        salvata comunque in suggested_category_name) -- finche' una voce non e'
        approvata dai 6 Giudici non deve comparire mescolata a voci gia' validate
        nella categoria di business reale (es. "IVA agricola"). Lo spostamento
        nella categoria reale avviene in erpv6_production/models/validation_session.py
        (action_human_approve) al momento dell'approvazione.

        :param entries: lista di dict {name, kb_type, category, content}
        :param source_label: valore per il campo 'source' di erpv6.kb
        :return: recordset erpv6.kb creati
        """
        category_model = self.env['erpv6.kb.category']
        kb_model = self.env['erpv6.kb']
        created = self.env['erpv6.kb']

        for entry in entries:
            # (name, kb_type) e' univoco a livello DB (erpv6_kb_name_unique).
            # Una voce puo' ripresentarsi identica tra blocchi diversi --
            # tipicamente riprendendo un'estrazione interrotta dopo che i
            # parametri di suddivisione blocchi sono cambiati (il checkpoint
            # kb_extraction_completed_chunks salvato non corrisponde piu' agli
            # stessi confini nella nuova suddivisione, quindi si puo' rientrare
            # su contenuto gia' estratto in precedenza). Senza questo controllo
            # l'INSERT falliva con IntegrityError e mandava in abort l'intera
            # transazione della sessione di estrazione, perdendo anche il
            # lavoro valido gia' fatto nello stesso batch.
            if kb_model.search_count([
                ('name', '=', entry['name']), ('kb_type', '=', entry['kb_type']),
            ]):
                _logger.warning(
                    "Estrazione KB: voce '%s' (kb_type=%s) gia' esistente, saltata "
                    "(probabile sovrapposizione con contenuto gia' estratto in "
                    "precedenza).", entry['name'], entry['kb_type'])
                continue
            staging_category = category_model.get_or_create(
                KB_STAGING_CATEGORY_NAME, entry['kb_type'], is_transversal=True)
            kb = kb_model.create({
                'name': entry['name'],
                'kb_type': entry['kb_type'],
                'category_id': staging_category.id,
                'suggested_category_name': entry['category'],
                'content': entry['content'],
                'content_format': 'text',
                'source': source_label or 'kb_extraction_service',
                # Mai attiva subito: deve passare da erpv6_validation (6 Giudici)
                # prima di essere raggiungibile da find_best_for() -- vedi
                # erpv6_production/models/validation_session.py, dove
                # action_human_approve() la riattiva su res_model='erpv6.kb'.
                'is_active': False,
                # Fase 1C.2 knowledge graph (vedi erpv6.kg.triple.shape):
                # gia' filtrate sul vocabolario controllato in _clean_triples,
                # mai scritte fuori da questo campo -- nessun nodo/arco reale
                # creato, solo dati che viaggiano con la voce KB stessa e che
                # i 6 Giudici vedranno via kb_validation_gate.
                'extracted_triples': entry.get('triples') or [],
            })
            created |= kb
        return created
